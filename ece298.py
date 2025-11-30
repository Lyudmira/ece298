# -*- coding: utf-8 -*-
"""
ECE-298 Lab B5 — 24h Operating Plan Solver (End-to-End)
======================================================

你给了 Table 1/2（泵曲线采样点 + 系统曲线 + 功率拟合式），这份脚本会：

1) 用“单调三次 Hermite (PCHIP)”对曲线做插值（比全局多项式拟合稳很多，不容易乱飙）。
2) 构建两个核心函数：
      get_flow(pump_rpm, zone_id) -> GPM   （泵曲线与系统曲线交点）
      get_power(pump_rpm) -> kW           （motor RPM = 2.5*pump RPM，再代入 Table2）
3) 计算每个 zone 的 Efficiency 曲线（Gallons/kWh）并画图。
4) 用优化（先连续 LP，再分钟级整数 ILP）生成 24h 排班：
      - 每个 zone 的供水：不超过需求（无浪费）
      - 每个 zone 的缺口：<= 100 gal
      - 总时长：覆盖完整 24h（若有空余会自动填 Idle=0kW）
      - 总成本：最小（TOU 分段电价）
5) 输出：
      - schedule_df（minute-resolution 的段表：Start/End/Zone/RPM/Flow/Power/Cost...）
      - summary（总成本/总电量/各 zone 误差）
      - 写 Excel（可选：填模板；否则生成新文件）

依赖（你说你会配环境）：numpy, pandas, scipy, matplotlib, openpyxl, pulp(推荐) 或 ortools(可选)
"""

from __future__ import annotations

import math
from dataclasses import dataclass
from datetime import datetime, timedelta
from functools import lru_cache
from typing import Dict, List, Tuple, Callable, Optional

import numpy as np
import pandas as pd

# ----------------------------
# 0) Baseline / Input Tables
# ----------------------------

GEAR_RATIO = 2.5  # Motor RPM = 2.5 * Pump RPM

DEMAND_GAL = {
    "inlet": 94_000,
    "zone1": 41_000,
    "zone2": 39_000,
    "zone3": 14_000,
}

# Time-of-Use periods for a 24h day (chronological)
# Off-Peak: 19:00-07:00, Mid: 07:00-11:00 & 17:00-19:00, On: 11:00-17:00
TOU_PERIODS = [
    # name, start_h, start_m, duration_hours, price($/kWh)
    ("off1", 0, 0, 7.0, 0.074),   # 00:00-07:00
    ("mid1", 7, 0, 4.0, 0.102),   # 07:00-11:00
    ("on", 11, 0, 6.0, 0.151),    # 11:00-17:00
    ("mid2", 17, 0, 2.0, 0.102),  # 17:00-19:00
    ("off2", 19, 0, 5.0, 0.074),  # 19:00-24:00
]
ZONE_ORDER = ["inlet", "zone3", "zone2", "zone1"]  # just for pretty scheduling order

# Table 1: Pump curves at discrete pump RPM. Points: (Flow GPM, Head ft)
PUMP_DATA: Dict[int, List[Tuple[float, float]]] = {
    100: [
        (0.34, 81.87),
        (38.51, 77.63),
        (71.98, 72.87),
        (96.08, 68.01),
        (120.68, 62.51),
        (139.28, 57.31),
        (163.88, 49.03),
        (170.15, 46.44),
        (184.99, 39.77),
        (193.07, 35.34),
        (202.47, 30.26),
    ],
    80: [
        (0.13, 67.60),
        (29.34, 64.59),
        (64.09, 59.43),
        (105.81, 50.87),
        (110.15, 49.44),
        (120.40, 46.79),
        (143.20, 39.48),
        (151.83, 36.45),
        (171.38, 27.62),
        (182.43, 21.65),
    ],
    60: [
        (0.13, 51.47),
        (28.49, 48.46),
        (57.70, 44.25),
        (86.06, 38.44),
        (98.54, 35.37),
        (123.56, 27.61),
        (143.84, 20.23),
        (155.57, 14.78),
    ],
    40: [
        (0.55, 32.93),
        (45.12, 27.75),
        (58.36, 25.30),
        (68.71, 23.60),
        (87.97, 19.23),
        (110.79, 13.30),
        (128.91, 7.39),
    ],
    20: [
        (0.55, 19.46),
        (17.61, 17.95),
        (37.52, 16.11),
        (43.91, 14.85),
        (55.45, 12.32),
        (71.56, 7.14),
        (79.45, 4.59),
    ],
}

# Table 1: System curves per zone: (Flow GPM, Head ft)
SYS_DATA: Dict[str, List[Tuple[float, float]]] = {
    "zone3": [  # blue(3)
        (0.00, 0.00),
        (9.08, 3.98),
        (18.68, 8.08),
        (30.41, 13.12),
        (37.52, 16.11),
        (45.33, 19.35),
        (58.36, 25.30),
        (71.34, 31.01),
        (86.06, 38.44),
        (97.78, 45.34),
        (105.81, 50.87),
        (120.68, 62.51),
        (128.70, 68.86),
        (134.24, 73.38),
    ],
    "zone2": [  # green(2)
        (0.00, 0.00),
        (12.71, 4.63),
        (17.40, 6.09),
        (25.29, 8.73),
        (36.16, 12.30),
        (43.91, 14.85),
        (53.01, 18.13),
        (68.71, 23.60),
        (79.23, 27.26),
        (90.53, 31.76),
        (98.54, 35.37),
        (107.59, 39.98),
        (120.40, 46.79),
        (133.18, 53.91),
        (139.28, 57.31),
        (143.41, 59.75),
    ],
    "inlet": [  # purple (INLET)
        (0.00, 0.00),
        (5.25, 1.45),
        (10.15, 2.51),
        (22.09, 5.27),
        (42.56, 9.49),
        (55.45, 12.32),
        (66.23, 14.76),
        (80.51, 17.65),
        (87.97, 19.23),
        (105.67, 23.18),
        (116.12, 25.82),
        (123.56, 27.61),
        (138.29, 31.50),
        (151.83, 36.45),
        (163.67, 42.23),
        (170.15, 46.44),
        (180.09, 56.05),
    ],
    "zone1": [  # red (1)
        (0.00, 0.00),
        (13.13, 0.90),
        (21.45, 1.54),
        (28.91, 2.46),
        (40.85, 3.49),
        (51.73, 4.80),
        (60.04, 5.71),
        (71.56, 7.14),
        (79.87, 8.45),
        (90.96, 10.02),
        (110.79, 13.30),
        (124.01, 15.94),
        (137.01, 18.57),
        (143.84, 20.23),
        (151.09, 22.00),
        (164.31, 25.43),
        (171.38, 27.62),
        (181.58, 30.98),
        (193.07, 35.34),
        (202.47, 39.46),
        (208.87, 42.64),
    ],
}

# Table 2: Motor power curve (kW) as function of motor RPM x
# y ≈ 7.2803e-6*(x+13.092)^(3.18385) for 0 <= x < 250.5
def motor_power_kw(motor_rpm: float) -> float:
    x = float(motor_rpm)
    return float(7.2803e-6 * ((x + 13.092) ** 3.18385))


# ----------------------------
# 1) PCHIP (Monotone Cubic Hermite) Interpolator
# ----------------------------

def _dedup_sorted_xy(x: np.ndarray, y: np.ndarray) -> Tuple[np.ndarray, np.ndarray]:
    """If x has duplicates, average their y's. Assumes x sorted."""
    ux, inv = np.unique(x, return_inverse=True)
    if len(ux) == len(x):
        return x, y
    y_acc = np.zeros_like(ux, dtype=float)
    counts = np.zeros_like(ux, dtype=int)
    for i, idx in enumerate(inv):
        y_acc[idx] += y[i]
        counts[idx] += 1
    return ux, (y_acc / counts)

def pchip_slopes(x: np.ndarray, y: np.ndarray) -> np.ndarray:
    """
    Compute PCHIP slopes (Fritsch–Carlson), preserving monotonicity.
    """
    x = np.asarray(x, dtype=float)
    y = np.asarray(y, dtype=float)
    n = len(x)
    if n < 2:
        raise ValueError("Need >=2 points")

    h = np.diff(x)
    delta = np.diff(y) / h
    d = np.zeros(n, dtype=float)

    # interior slopes
    for k in range(1, n - 1):
        if delta[k - 1] == 0.0 or delta[k] == 0.0 or np.sign(delta[k - 1]) != np.sign(delta[k]):
            d[k] = 0.0
        else:
            w1 = 2 * h[k] + h[k - 1]
            w2 = h[k] + 2 * h[k - 1]
            d[k] = (w1 + w2) / (w1 / delta[k - 1] + w2 / delta[k])

    # endpoints
    if n == 2:
        d[0] = d[1] = delta[0]
        return d

    d0 = ((2 * h[0] + h[1]) * delta[0] - h[0] * delta[1]) / (h[0] + h[1])
    if np.sign(d0) != np.sign(delta[0]):
        d0 = 0.0
    elif np.sign(delta[0]) != np.sign(delta[1]) and abs(d0) > abs(3 * delta[0]):
        d0 = 3 * delta[0]
    d[0] = d0

    dn = ((2 * h[-1] + h[-2]) * delta[-1] - h[-1] * delta[-2]) / (h[-1] + h[-2])
    if np.sign(dn) != np.sign(delta[-1]):
        dn = 0.0
    elif np.sign(delta[-1]) != np.sign(delta[-2]) and abs(dn) > abs(3 * delta[-1]):
        dn = 3 * delta[-1]
    d[-1] = dn

    return d

def pchip_interpolator(x: List[float], y: List[float]) -> Callable[[np.ndarray], np.ndarray]:
    """
    Return a function f(xq) that evaluates PCHIP interpolation at xq.
    Extrapolation is *clamped* to the boundary interval.
    """
    x = np.asarray(x, dtype=float)
    y = np.asarray(y, dtype=float)

    order = np.argsort(x)
    x = x[order]
    y = y[order]
    x, y = _dedup_sorted_xy(x, y)

    d = pchip_slopes(x, y)

    def f(xq):
        xq = np.asarray(xq, dtype=float)
        # Locate interval
        idx = np.searchsorted(x, xq) - 1
        idx = np.clip(idx, 0, len(x) - 2)

        x0 = x[idx]
        x1 = x[idx + 1]
        y0 = y[idx]
        y1 = y[idx + 1]
        d0 = d[idx]
        d1 = d[idx + 1]
        h = (x1 - x0)
        # avoid divide-by-zero (shouldn't happen after dedup)
        t = (xq - x0) / h

        # Hermite basis
        h00 = (2 * t**3 - 3 * t**2 + 1)
        h10 = (t**3 - 2 * t**2 + t)
        h01 = (-2 * t**3 + 3 * t**2)
        h11 = (t**3 - t**2)
        return h00 * y0 + h10 * h * d0 + h01 * y1 + h11 * h * d1

    return f


# ----------------------------
# 2) Pump-System model: get_flow & get_power
# ----------------------------

def lerp(x, x0, x1, y0, y1):
    if x1 == x0:
        return y0
    return y0 + (x - x0) / (x1 - x0) * (y1 - y0)

@dataclass(frozen=True)
class Period:
    name: str
    start: datetime
    end: datetime
    price: float  # $/kWh

    @property
    def hours(self) -> float:
        return (self.end - self.start).total_seconds() / 3600.0

    @property
    def minutes(self) -> int:
        return int(round((self.end - self.start).total_seconds() / 60.0))

class PumpSystem:
    """
    Builds:
      - pump head curve family based on discrete RPM curves.
      - system head curves for each zone.
      - get_flow(rpm, zone) via intersection (bisection).
      - get_power(rpm) from Table2 + gear ratio.
    """

    def __init__(self, pump_data: Dict[int, List[Tuple[float, float]]], sys_data: Dict[str, List[Tuple[float, float]]]):
        self.pump_data = {int(k): list(v) for k, v in pump_data.items()}
        self.sys_data = {str(k): list(v) for k, v in sys_data.items()}

        self.measured_rpms = sorted(self.pump_data.keys())

        # Build normalized pump curves: head vs s (= Q / Qmax) using PCHIP
        self.pump_qmax = {}
        self.pump_head_s_func = {}
        for rpm, pts in self.pump_data.items():
            q = np.array([p[0] for p in pts], dtype=float)
            h = np.array([p[1] for p in pts], dtype=float)
            qmax = float(np.max(q))
            self.pump_qmax[rpm] = qmax

            # add an explicit q=0 point if missing (use first head as "shutoff-ish" proxy)
            # helps bracketing near zero-flow.
            if q.min() > 1e-9:
                q = np.insert(q, 0, 0.0)
                h = np.insert(h, 0, h[0])

            s = q / qmax
            self.pump_head_s_func[rpm] = pchip_interpolator(list(s), list(h))

        # Build system head funcs: head vs Q using PCHIP
        self.sys_qmax = {}
        self.sys_head_func = {}
        for zone, pts in self.sys_data.items():
            q = np.array([p[0] for p in pts], dtype=float)
            h = np.array([p[1] for p in pts], dtype=float)
            self.sys_qmax[zone] = float(np.max(q))
            self.sys_head_func[zone] = pchip_interpolator(list(q), list(h))

    def get_power_kw(self, pump_rpm: float) -> float:
        motor_rpm = GEAR_RATIO * float(pump_rpm)
        return motor_power_kw(motor_rpm)

    def _pump_qmax_at(self, pump_rpm: float) -> float:
        r = float(pump_rpm)
        if r <= self.measured_rpms[0]:
            return self.pump_qmax[self.measured_rpms[0]]
        if r >= self.measured_rpms[-1]:
            return self.pump_qmax[self.measured_rpms[-1]]
        for i in range(len(self.measured_rpms) - 1):
            r0, r1 = self.measured_rpms[i], self.measured_rpms[i + 1]
            if r0 <= r <= r1:
                return float(lerp(r, r0, r1, self.pump_qmax[r0], self.pump_qmax[r1]))
        return self.pump_qmax[self.measured_rpms[-1]]

    def _pump_head(self, pump_rpm: float, flow_gpm: float) -> float:
        """
        Interpolate head at given (pump_rpm, flow_gpm).
        Strategy:
          - find bounding rpm curves r0,r1
          - evaluate each curve at its own s = Q/Qmax(r_i)
          - linearly interpolate head across rpm
        """
        r = float(pump_rpm)
        q = max(0.0, float(flow_gpm))

        # clamp rpm to measured range for interpolation
        if r <= self.measured_rpms[0]:
            r0 = r1 = self.measured_rpms[0]
        elif r >= self.measured_rpms[-1]:
            r0 = r1 = self.measured_rpms[-1]
        else:
            r0 = r1 = None
            for i in range(len(self.measured_rpms) - 1):
                a, b = self.measured_rpms[i], self.measured_rpms[i + 1]
                if a <= r <= b:
                    r0, r1 = a, b
                    break
            assert r0 is not None and r1 is not None

        qmax0 = self.pump_qmax[r0]
        qmax1 = self.pump_qmax[r1]

        # avoid s>1: clamp (endpoints are "definition domain ends" per your note)
        s0 = min(1.0, q / qmax0) if qmax0 > 0 else 0.0
        s1 = min(1.0, q / qmax1) if qmax1 > 0 else 0.0

        h0 = float(self.pump_head_s_func[r0](np.array([s0]))[0])
        h1 = float(self.pump_head_s_func[r1](np.array([s1]))[0])
        if r0 == r1:
            return h0
        return float(lerp(r, r0, r1, h0, h1))

    def _sys_head(self, zone: str, flow_gpm: float) -> float:
        q = max(0.0, float(flow_gpm))
        qmax = self.sys_qmax[zone]
        if q > qmax:
            # Definition domain ends at qmax; beyond that system curve isn't defined.
            # We'll treat as NaN so solver doesn't wander there.
            return float("nan")
        return float(self.sys_head_func[zone](np.array([q]))[0])

    @lru_cache(maxsize=20000)
    def get_flow_gpm(self, pump_rpm: float, zone: str) -> float:
        """
        Solve operating point where pump_head(r,Q) = sys_head(zone,Q).
        Return Q in GPM. If can't overcome system at near-zero flow -> 0.
        """
        zone = str(zone)
        r = float(pump_rpm)

        # cap Q search to the minimum of pump domain and system domain
        qmax = min(self._pump_qmax_at(r), self.sys_qmax[zone])
        if qmax <= 1e-9:
            return 0.0

        def g(q):
            hp = self._pump_head(r, q)
            hs = self._sys_head(zone, q)
            if math.isnan(hs):
                return float("nan")
            return hp - hs

        # Evaluate near zero to determine feasibility
        qeps = min(0.1, qmax * 1e-4 + 1e-6)
        glo = g(qeps)
        if math.isnan(glo):
            qeps = 1e-3
            glo = g(qeps)

        if math.isnan(glo):
            return 0.0  # give up
        if glo < 0:
            return 0.0  # can't push water uphill even at (almost) zero flow

        # Evaluate at upper end
        ghi = g(qmax)
        if math.isnan(ghi):
            # back off a bit
            for frac in [0.999, 0.99, 0.95, 0.9, 0.8]:
                qq = qmax * frac
                val = g(qq)
                if not math.isnan(val):
                    qmax = qq
                    ghi = val
                    break

        if math.isnan(ghi):
            return 0.0
        if ghi > 0:
            # pump curve above system curve even at max defined flow => operating point beyond domain; clamp
            return float(qmax)

        # Bisection between qeps and qmax where sign changes
        a, b = qeps, qmax
        fa, fb = glo, ghi
        for _ in range(70):
            m = 0.5 * (a + b)
            fm = g(m)
            if math.isnan(fm):
                b = m
                fb = fm
                continue
            if abs(fm) < 1e-9 or (b - a) < 1e-7:
                return float(m)
            if np.sign(fm) == np.sign(fa):
                a, fa = m, fm
            else:
                b, fb = m, fm
        return float(0.5 * (a + b))

    def efficiency_gal_per_kwh(self, pump_rpm: float, zone: str) -> float:
        q = self.get_flow_gpm(pump_rpm, zone)
        p = self.get_power_kw(pump_rpm)
        if p <= 0 or q <= 0:
            return float("nan")
        return (q * 60.0) / p

    def hours_for_volume(self, pump_rpm: float, zone: str, volume_gal: float) -> float:
        q = self.get_flow_gpm(pump_rpm, zone)
        if q <= 0:
            return float("inf")
        return float(volume_gal) / (q * 60.0)


# ----------------------------
# 3) Analysis: Sweet spot plots + quick checks
# ----------------------------

def sweet_spot_analysis(model: PumpSystem, rpm_min=20, rpm_max=100, out_png: Optional[str] = "efficiency.png"):
    import matplotlib.pyplot as plt

    rpms = np.arange(rpm_min, rpm_max + 1, 1)
    zones = list(DEMAND_GAL.keys())

    sweet = {}
    plt.figure()
    for z in zones:
        eff = np.array([model.efficiency_gal_per_kwh(r, z) for r in rpms], dtype=float)
        # ignore nan
        idx = np.nanargmax(eff)
        sweet[z] = {
            "rpm": int(rpms[idx]),
            "eff_gal_per_kwh": float(eff[idx]),
            "flow_gpm": float(model.get_flow_gpm(int(rpms[idx]), z)),
            "power_kw": float(model.get_power_kw(int(rpms[idx]))),
        }
        plt.plot(rpms, eff, label=f"{z}")

    plt.xlabel("Pump RPM")
    plt.ylabel("Efficiency (gal/kWh) = Q*60 / P")
    plt.title("RPM vs Efficiency (gal/kWh)")
    plt.legend()
    plt.grid(True, alpha=0.2)

    if out_png:
        plt.savefig(out_png, dpi=200, bbox_inches="tight")

    return sweet


# ----------------------------
# 4) Optimization
#    Step A: Continuous LP (SciPy linprog) -> good coarse optimum
#    Step B: Minute-level Integer ILP (PuLP) -> schedule-ready + tolerance
# ----------------------------

def build_periods(base_date: datetime) -> List[Period]:
    periods = []
    for name, sh, sm, dur_h, price in TOU_PERIODS:
        start = base_date.replace(hour=sh, minute=sm, second=0, microsecond=0)
        end = start + timedelta(hours=dur_h)
        periods.append(Period(name=name, start=start, end=end, price=price))
    return periods

def solve_continuous_lp(model: PumpSystem, periods: List[Period], rpm_min=20, rpm_max=100):
    """
    Continuous LP:
      variables t[p,z,r] = hours using zone z at rpm r during period p.
      minimize sum t * power(r) * price(p)
      subject to:
         sum_{z,r} t[p,z,r] <= period_hours[p]
         sum_{p,r} t[p,z,r]*flow(z,r)*60 == demand[z]
    """
    try:
        from scipy.optimize import linprog
    except Exception as e:
        raise RuntimeError("Need scipy.optimize.linprog for the continuous LP stage.") from e

    zones = list(DEMAND_GAL.keys())
    rpms = list(range(rpm_min, rpm_max + 1))

    # Precompute tables
    flow = {(z, r): model.get_flow_gpm(r, z) for z in zones for r in rpms}
    power = {r: model.get_power_kw(r) for r in rpms}

    # Build variable index
    var = []
    for p in periods:
        for z in zones:
            for r in rpms:
                # skip useless decisions
                if flow[(z, r)] <= 1e-9:
                    continue
                var.append((p.name, z, r))
    n = len(var)

    # objective coefficients: $/hour = kW * $/kWh
    c = np.zeros(n, dtype=float)
    for j, (pn, z, r) in enumerate(var):
        price = next(pp.price for pp in periods if pp.name == pn)
        c[j] = power[r] * price

    # period time constraints
    A_ub = np.zeros((len(periods), n), dtype=float)
    b_ub = np.array([p.hours for p in periods], dtype=float)
    for i, p in enumerate(periods):
        for j, (pn, z, r) in enumerate(var):
            if pn == p.name:
                A_ub[i, j] = 1.0

    # zone volume equality constraints
    A_eq = np.zeros((len(zones), n), dtype=float)
    b_eq = np.array([DEMAND_GAL[z] for z in zones], dtype=float)
    for i, z in enumerate(zones):
        for j, (pn, zz, r) in enumerate(var):
            if zz == z:
                A_eq[i, j] = flow[(z, r)] * 60.0  # gal/hour

    bounds = [(0, None)] * n
    res = linprog(c, A_ub=A_ub, b_ub=b_ub, A_eq=A_eq, b_eq=b_eq, bounds=bounds, method="highs")
    if res.status != 0:
        raise RuntimeError(f"Continuous LP failed: {res.message}")

    # Extract sparse solution
    x = res.x
    sol = [(var[i][0], var[i][1], var[i][2], float(x[i])) for i in range(n) if x[i] > 1e-8]
    sol.sort(key=lambda t: (t[0], ZONE_ORDER.index(t[1]), t[2]))
    return {
        "objective_cost": float(res.fun),
        "solution_hours": sol,
        "flow_table": flow,
        "power_table": power,
    }

def minute_ilp_refine(
    model: PumpSystem,
    periods: List[Period],
    coarse_solution_hours: List[Tuple[str, str, int, float]],
    rpm_min=20,
    rpm_max=100,
    tol_gal=100.0,
    include_idle=True,
    penalty_shortfall_per_gal=0.02,  # $/gal penalty to prefer meeting demand over "cheap underdelivery"
):
    """
    Minute-level ILP:
      m[p,z,r] = integer minutes allocated to zone z at rpm r in period p.
      objective = energy_cost + penalty * shortfall_gal
      constraints:
        - sum_{z,r} m[p,z,r] == period_minutes[p] (if include_idle True, idle can absorb)
        - delivered_z + shortfall_z == demand_z
            where 0 <= shortfall_z <= tol_gal
        - delivered_z <= demand_z (no oversupply enforced by shortfall definition)
    Candidate RPM reduction:
      Use only RPMs that appear in coarse LP +/- 2 (per zone), to keep ILP small & stable.
    """
    try:
        import pulp
    except Exception as e:
        raise RuntimeError(
            "Minute ILP stage requires PuLP. Install e.g. `pip install pulp` "
            "and ensure a MILP solver is available (CBC recommended)."
        ) from e

    zones = list(DEMAND_GAL.keys())
    idle_zone = "idle"

    period_by_name = {p.name: p for p in periods}

    # Candidate rpm sets per zone from coarse solution (+/-2)
    cand = {z: set() for z in zones}
    for pn, z, r, hrs in coarse_solution_hours:
        cand[z].add(int(r))
        for dr in [-2, -1, 1, 2]:
            rr = int(r) + dr
            if rpm_min <= rr <= rpm_max:
                cand[z].add(rr)

    # Ensure each zone has at least some rpms (fallback: add a small spread)
    for z in zones:
        if not cand[z]:
            cand[z] = set(range(rpm_min, rpm_max + 1))
        # remove rpms that yield zero flow to that zone
        cand[z] = {r for r in cand[z] if model.get_flow_gpm(r, z) > 1e-9}

        # still empty? then it's physically impossible in given range
        if not cand[z]:
            raise RuntimeError(f"No feasible RPM in [{rpm_min},{rpm_max}] produces positive flow for {z}.")

    # Build MILP
    prob = pulp.LpProblem("Plan_Minute_ILP", pulp.LpMinimize)

    # decision variables: minutes
    m = {}
    for p in periods:
        for z in zones:
            for r in sorted(cand[z]):
                m[(p.name, z, r)] = pulp.LpVariable(f"m_{p.name}_{z}_{r}", lowBound=0, cat=pulp.LpInteger)

        if include_idle:
            m[(p.name, idle_zone, 0)] = pulp.LpVariable(f"m_{p.name}_{idle_zone}_0", lowBound=0, cat=pulp.LpInteger)

    # Shortfall variables (gal)
    shortfall = {z: pulp.LpVariable(f"short_{z}", lowBound=0, upBound=tol_gal, cat=pulp.LpContinuous) for z in zones}

    # Period minutes must match exactly (no blank)
    for p in periods:
        expr = []
        for z in zones:
            expr += [m[(p.name, z, r)] for r in cand[z]]
        if include_idle:
            expr += [m[(p.name, idle_zone, 0)]]
        prob += (pulp.lpSum(expr) == p.minutes), f"PeriodMinutes_{p.name}"

    # Zone delivery constraints
    for z in zones:
        delivered = []
        for p in periods:
            for r in cand[z]:
                delivered.append(m[(p.name, z, r)] * model.get_flow_gpm(r, z))  # gal/min
        prob += (pulp.lpSum(delivered) + shortfall[z] == DEMAND_GAL[z]), f"DemandEq_{z}"
        # By construction, delivered <= demand because shortfall >=0 (so no oversupply)

    # Objective: cost + penalty * shortfall
    cost_terms = []
    for p in periods:
        price = p.price
        for z in zones:
            for r in cand[z]:
                power_kw = model.get_power_kw(r)
                # cost per minute = (kW * (1/60) h) * $/kWh
                cost_terms.append(m[(p.name, z, r)] * (power_kw / 60.0) * price)
        if include_idle:
            # idle has 0 cost; skip
            pass

    penalty_terms = [shortfall[z] * penalty_shortfall_per_gal for z in zones]
    prob += pulp.lpSum(cost_terms) + pulp.lpSum(penalty_terms)

    # Solve
    # CBC is usually available via pulp; user says they can configure environment.
    prob.solve(pulp.PULP_CBC_CMD(msg=False))

    status = pulp.LpStatus[prob.status]
    if status != "Optimal":
        raise RuntimeError(f"Minute ILP failed. Status={status}")

    # Extract minute allocations
    allocations = []
    for (pn, z, r), var in m.items():
        val = int(round(pulp.value(var)))
        if val > 0:
            allocations.append((pn, z, r, val))

    # Extract shortfalls
    shortfalls = {z: float(pulp.value(shortfall[z])) for z in zones}

    # Compute true cost w/o penalty
    true_cost = 0.0
    for pn, z, r, mins in allocations:
        if z == idle_zone:
            continue
        p = period_by_name[pn]
        true_cost += (model.get_power_kw(r) * (mins / 60.0) * p.price)

    return {
        "allocations_minutes": allocations,
        "shortfalls": shortfalls,
        "true_cost": true_cost,
        "candidate_rpms": {z: sorted(cand[z]) for z in zones},
    }


# ----------------------------
# 5) Build a chronological schedule + compute totals
# ----------------------------

def allocations_to_schedule_df(model: PumpSystem, periods: List[Period], allocations_minutes: List[Tuple[str, str, int, int]]):
    """
    Convert (period_name, zone, rpm, minutes) into chronological segments with Start/End times.
    Within each period, we order segments by (zone priority, rpm desc), then compress adjacent identical segments.
    """
    period_by_name = {p.name: p for p in periods}

    # group by period
    grouped: Dict[str, List[Tuple[str, int, int]]] = {p.name: [] for p in periods}
    for pn, z, r, mins in allocations_minutes:
        grouped[pn].append((z, r, mins))

    # stable sort for readability
    zprio = {z: i for i, z in enumerate(ZONE_ORDER + ["idle"])}
    for pn in grouped:
        grouped[pn].sort(key=lambda t: (zprio.get(t[0], 999), -t[1]))

    # Expand to chronological segments with timestamps
    rows = []
    base = periods[0].start.replace(hour=0, minute=0, second=0, microsecond=0)
    # Ensure base day alignment
    for p in periods:
        t = p.start
        for z, r, mins in grouped[p.name]:
            if mins <= 0:
                continue
            end = t + timedelta(minutes=mins)
            rows.append((t, end, p.name, z, r, mins))
            t = end
        # Sanity: period fully covered
        if abs((t - p.end).total_seconds()) > 1e-6:
            raise RuntimeError(f"Period {p.name} not fully covered: ended at {t}, expected {p.end}")

    # Compress adjacent segments if identical zone/rpm and contiguous time
    compressed = []
    for seg in rows:
        if not compressed:
            compressed.append(list(seg))
            continue
        prev = compressed[-1]
        if prev[1] == seg[0] and prev[3] == seg[3] and prev[4] == seg[4] and prev[2] == seg[2]:
            # same period, zone, rpm: merge
            prev[1] = seg[1]
            prev[5] += seg[5]
        else:
            compressed.append(list(seg))

    # Build DataFrame with computed flow/power/volume/energy/cost
    out = []
    for start, end, tou, zone, rpm, mins in compressed:
        if zone == "idle":
            flow = 0.0
            power = 0.0
            motor_rpm = 0.0
        else:
            flow = model.get_flow_gpm(rpm, zone)
            power = model.get_power_kw(rpm)
            motor_rpm = GEAR_RATIO * rpm

        hours = mins / 60.0
        vol = flow * mins  # gal = (gal/min)*min
        energy = power * hours  # kWh
        price = next(pp.price for pp in periods if pp.name == tou)
        cost = energy * price

        out.append({
            "Start": start.strftime("%H:%M"),
            "End": end.strftime("%H:%M"),
            "TOU": tou,
            "Zone": zone,
            "Pump RPM": rpm,
            "Motor RPM": motor_rpm,
            "Flow (GPM)": flow,
            "Power (kW)": power,
            "Minutes": mins,
            "Volume (gal)": vol,
            "Energy (kWh)": energy,
            "Cost ($)": cost,
        })

    df = pd.DataFrame(out)
    return df

def compute_summary(df: pd.DataFrame):
    # delivered by zone (excluding idle)
    delivered = df.groupby("Zone")["Volume (gal)"].sum().to_dict()
    cost_total = float(df["Cost ($)"].sum())
    energy_total = float(df["Energy (kWh)"].sum())

    # Only evaluate demand zones
    errors = {}
    for z, demand in DEMAND_GAL.items():
        got = float(delivered.get(z, 0.0))
        errors[z] = got - demand  # should be <=0, and >= -100 ideally

    # TOU breakdown
    cost_by_tou = df.groupby("TOU")["Cost ($)"].sum().to_dict()
    energy_by_tou = df.groupby("TOU")["Energy (kWh)"].sum().to_dict()

    return {
        "delivered_gal": delivered,
        "errors_gal": errors,
        "total_cost": cost_total,
        "total_energy_kwh": energy_total,
        "cost_by_tou": cost_by_tou,
        "energy_by_tou": energy_by_tou,
    }


# ----------------------------
# 6) Excel output (template-aware best effort)
# ----------------------------

def write_excel(schedule_df: pd.DataFrame, summary: dict, out_path: str, template_path: Optional[str] = None):
    """
    If template_path is provided:
      - try to open and fill a sheet containing headers like Start/End/Zone/Pump RPM.
      - otherwise, add a new sheet named 'OperatingPlan'.
    If no template:
      - create a new workbook with sheets: OperatingPlan, Summary.
    """
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils import get_column_letter

    def autofit(ws):
        # simple width sizing
        for col in range(1, ws.max_column + 1):
            maxlen = 0
            for row in range(1, ws.max_row + 1):
                v = ws.cell(row=row, column=col).value
                if v is None:
                    continue
                maxlen = max(maxlen, len(str(v)))
            ws.column_dimensions[get_column_letter(col)].width = min(45, max(10, maxlen + 2))

    if template_path:
        wb = load_workbook(template_path)
        ws = wb.active

        # search header row
        wanted = ["Start", "End", "TOU", "Zone", "Pump RPM", "Motor RPM", "Flow (GPM)", "Power (kW)", "Minutes", "Volume (gal)", "Energy (kWh)", "Cost ($)"]
        header_row = None
        header_map = {}

        for r in range(1, min(60, ws.max_row + 1)):
            row_vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            if not any(row_vals):
                continue
            # map cell value -> column index
            m = {str(v).strip(): i + 1 for i, v in enumerate(row_vals) if v is not None}
            if "Start" in m and "End" in m and "Zone" in m:
                header_row = r
                header_map = m
                break

        if header_row is None:
            # create new sheet
            ws = wb.create_sheet("OperatingPlan")
            header_row = 1
            header_map = {h: i + 1 for i, h in enumerate(wanted)}
            for i, h in enumerate(wanted, 1):
                ws.cell(header_row, i).value = h

        # write data below header
        start_row = header_row + 1
        # clear old rows (best-effort)
        for r in range(start_row, start_row + 500):
            for h in wanted:
                c = header_map.get(h)
                if c:
                    ws.cell(r, c).value = None

        for i, row in enumerate(schedule_df.to_dict(orient="records")):
            rr = start_row + i
            for h in wanted:
                cc = header_map.get(h)
                if not cc:
                    continue
                ws.cell(rr, cc).value = row.get(h)

        # add Summary sheet
        if "Summary" in wb.sheetnames:
            ws2 = wb["Summary"]
            ws2.delete_rows(1, ws2.max_row)
        else:
            ws2 = wb.create_sheet("Summary")

        ws2["A1"].value = "Metric"
        ws2["B1"].value = "Value"
        r = 2
        for k, v in summary.items():
            ws2.cell(r, 1).value = str(k)
            ws2.cell(r, 2).value = str(v)
            r += 1

        autofit(ws)
        autofit(ws2)
        wb.save(out_path)
        return

    # no template -> new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "OperatingPlan"

    for j, col in enumerate(schedule_df.columns, 1):
        ws.cell(1, j).value = col
    for i, row in enumerate(schedule_df.itertuples(index=False), 2):
        for j, v in enumerate(row, 1):
            ws.cell(i, j).value = v

    ws2 = wb.create_sheet("Summary")
    ws2["A1"].value = "Metric"
    ws2["B1"].value = "Value"
    r = 2
    for k, v in summary.items():
        ws2.cell(r, 1).value = str(k)
        ws2.cell(r, 2).value = str(v)
        r += 1

    autofit(ws)
    autofit(ws2)
    wb.save(out_path)


# ----------------------------
# 7) Main runner
# ----------------------------

def main(
    rpm_min=20,
    rpm_max=100,
    tol_gal=100.0,
    out_excel="operating_plan_output.xlsx",
    template_path: Optional[str] = None,
    out_eff_png="efficiency.png",
):
    # build model
    model = PumpSystem(PUMP_DATA, SYS_DATA)

    # sweet spot analysis + plot
    sweet = sweet_spot_analysis(model, rpm_min=rpm_min, rpm_max=rpm_max, out_png=out_eff_png)
    print("\n=== Sweet Spots (max gal/kWh) ===")
    for z, info in sweet.items():
        print(f"{z:>5}: rpm={info['rpm']}, eff={info['eff_gal_per_kwh']:.3f} gal/kWh, "
              f"Q={info['flow_gpm']:.3f} GPM, P={info['power_kw']:.3f} kW")

    # periods
    base_date = datetime(2025, 1, 1, 0, 0, 0)  # arbitrary, only HH:MM matters
    periods = build_periods(base_date)

    # continuous LP (coarse)
    lp = solve_continuous_lp(model, periods, rpm_min=rpm_min, rpm_max=rpm_max)
    print("\n=== Continuous LP (coarse) ===")
    print(f"LP objective cost (continuous, $): {lp['objective_cost']:.6f}")
    print("Non-zero decisions (period, zone, rpm, hours):")
    for pn, z, r, hrs in lp["solution_hours"]:
        print(f"  {pn:>4}  {z:>5}  rpm={r:>3}  hours={hrs:.6f}")

    # minute ILP refine
    ilp = minute_ilp_refine(
        model,
        periods,
        lp["solution_hours"],
        rpm_min=rpm_min,
        rpm_max=rpm_max,
        tol_gal=tol_gal,
        include_idle=True,
        penalty_shortfall_per_gal=0.02,  # pushes towards meeting demand unless impossible
    )
    print("\n=== Minute ILP (final schedule-ready) ===")
    print(f"True cost (no shortfall penalty) = ${ilp['true_cost']:.6f}")
    print("Shortfalls (gal):", ilp["shortfalls"])
    print("Candidate RPMs by zone:", ilp["candidate_rpms"])

    # build schedule df
    schedule_df = allocations_to_schedule_df(model, periods, ilp["allocations_minutes"])
    summary = compute_summary(schedule_df)

    print("\n=== Final Totals ===")
    print("Total cost ($):", summary["total_cost"])
    print("Total energy (kWh):", summary["total_energy_kwh"])
    print("Delivered (gal) by zone:", {k: float(v) for k, v in summary["delivered_gal"].items() if k != "idle"})
    print("Errors (gal) by zone (delivered - demand):", summary["errors_gal"])
    print("Cost by TOU:", summary["cost_by_tou"])
    print("Energy by TOU:", summary["energy_by_tou"])

    # enforce rubric constraints (no oversupply; abs error <= 100)
    for z, err in summary["errors_gal"].items():
        if err > 1e-6:
            raise RuntimeError(f"Oversupply detected for {z}: {err} gal (must be <= 0)")
        if abs(err) - 1e-6 > tol_gal:
            raise RuntimeError(f"Error too large for {z}: {err} gal (|err| must be <= {tol_gal})")

    # write Excel
    write_excel(schedule_df, summary, out_path=out_excel, template_path=template_path)
    print(f"\nWrote Excel: {out_excel}")
    print(f"Saved efficiency plot: {out_eff_png}")

    return schedule_df, summary


if __name__ == "__main__":
    # 你可以在这里改输出路径/模板路径
    # template_path = "YOUR_TEMPLATE.xlsx"
    template_path = None
    main(
        rpm_min=20,
        rpm_max=100,
        tol_gal=100.0,
        out_excel="operating_plan_output.xlsx",
        template_path=template_path,
        out_eff_png="efficiency.png",
    )
